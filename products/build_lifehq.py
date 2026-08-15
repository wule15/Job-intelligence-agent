import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()

# ---------- palette (shared brand) ----------
ARIAL = "Arial"
HDR_FILL   = PatternFill("solid", fgColor="1F2937")
HDR_FONT   = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
INPUT_FILL = PatternFill("solid", fgColor="EAF2FB")
CALC_FILL  = PatternFill("solid", fgColor="F3F4F6")
CARD_FILL  = PatternFill("solid", fgColor="D1FAE5")   # soft green = money/life accent
WARN_FILL  = PatternFill("solid", fgColor="FEE2E2")
TITLE_FONT = Font(name=ARIAL, bold=True, size=18, color="111827")
SUB_FONT   = Font(name=ARIAL, size=11, color="4B5563")
BOLD = Font(name=ARIAL, bold=True, size=11, color="111827")
BODY = Font(name=ARIAL, size=11, color="111827")
ACCENT = Font(name=ARIAL, bold=True, size=12, color="059669")
BIGNUM = Font(name=ARIAL, bold=True, size=16, color="059669")
thin = Side(style="thin", color="D1D5DB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = "#,##0.00"; PCT = "0%"

def hdr(ws, row, headers, start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start+i, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

def inp(cell):
    cell.fill = INPUT_FILL; cell.font = BODY; cell.border = BORDER
    cell.alignment = Alignment(horizontal="center")

categories = ["Rent / Mortgage","Utilities","Groceries","Dining Out","Transport","Fuel",
              "Insurance","Subscriptions","Shopping","Health","Entertainment","Savings","Other"]
habits = ["Exercise","Read 10 pages","Drink water","No junk food","Sleep by 11","Study","Meditate","Steps 8k"]
EXP_N = 201  # expenses rows 2..201

# ================= READ ME =================
ws = wb.active; ws.title = "Read Me"; ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2; ws.column_dimensions["B"].width = 100
ws["B2"] = "LIFE HQ"; ws["B2"].font = TITLE_FONT
ws["B3"] = "Your money and your habits, on one dashboard."; ws["B3"].font = SUB_FONT
lines = [
    ("", ""),
    ("HOW IT WORKS", "h"),
    ("1.  Set your monthly income and a budget per category on the 'Budget' tab (blue cells).", "b"),
    ("2.  Log spending on the 'Expenses' tab — pick a category from the dropdown, type the amount.", "b"),
    ("3.  Tick off habits each day on the 'Habits' tab (type an 'x' or anything in the day cell).", "b"),
    ("4.  The 'Dashboard' tab pulls it all together — budget used, savings progress, habit completion. Nothing to fill in there.", "b"),
    ("", ""),
    ("WHAT IT TRACKS FOR YOU", "h"),
    ("•  Budget vs. actual per category, and how much of your income is left.", "b"),
    ("•  Savings goals with automatic % complete.", "b"),
    ("•  Habit completion for the month, your best habit, and total check-ins.", "b"),
    ("•  Your top spending categories, ranked automatically.", "b"),
    ("", ""),
    ("TIPS", "h"),
    ("•  Edit categories and habits on the 'Setup' tab — the budget rows, dropdowns and habit grid follow.", "b"),
    ("•  No currency symbol is forced, so it works in any currency — just stay consistent.", "b"),
    ("•  New month? Right-click the 'Expenses' and 'Habits' tabs and Duplicate them to start fresh, or clear the cells.", "b"),
    ("", ""),
    ("LEGEND", "h"),
    ("BLUE cell = you fill this        GREY cell = calculated for you, don't edit", "b"),
]
r = 5
for text, kind in lines:
    c = ws.cell(row=r, column=2, value=text); c.font = ACCENT if kind == "h" else BODY; r += 1

# ================= SETUP =================
st = wb.create_sheet("Setup"); st.sheet_view.showGridLines = False
st.column_dimensions["A"].width = 22; st.column_dimensions["B"].width = 20
hdr(st, 1, ["Expense Categories", "Habits"])
for i in range(20):
    a = st.cell(row=2+i, column=1, value=(categories[i] if i < len(categories) else None)); inp(a); a.alignment = Alignment(horizontal="left")
    b = st.cell(row=2+i, column=2, value=(habits[i] if i < len(habits) else None)); inp(b); b.alignment = Alignment(horizontal="left")

# ================= EXPENSES =================
ex = wb.create_sheet("Expenses"); ex.sheet_view.showGridLines = False
for col, w in zip("ABCD", [13, 20, 30, 12]):
    ex.column_dimensions[col].width = w
hdr(ex, 1, ["Date", "Category", "Description", "Amount"]); ex.freeze_panes = "A2"
sample_exp = [
    (date(2026,8,1),"Rent / Mortgage","Monthly rent",650),
    (date(2026,8,2),"Groceries","Weekly shop",70),
    (date(2026,8,3),"Subscriptions","Streaming",15),
    (date(2026,8,4),"Dining Out","Dinner w/ friends",35),
    (date(2026,8,5),"Fuel","Full tank",60),
    (date(2026,8,6),"Groceries","Mid-week top up",40),
    (date(2026,8,7),"Utilities","Electric + water",90),
    (date(2026,8,9),"Entertainment","Cinema",18),
    (date(2026,8,10),"Shopping","New shoes",80),
    (date(2026,8,12),"Health","Gym membership",30),
    (date(2026,8,14),"Dining Out","Lunch",14),
    (date(2026,8,15),"Transport","Bus pass",25),
]
for i in range(EXP_N - 1):
    r = 2 + i
    for col in range(1, 5):
        c = ex.cell(row=r, column=col); inp(c)
    ex.cell(row=r, column=1).number_format = "yyyy-mm-dd"
    ex.cell(row=r, column=4).number_format = MONEY
    if i < len(sample_exp):
        d, cat, desc, amt = sample_exp[i]
        ex.cell(row=r, column=1, value=d); ex.cell(row=r, column=2, value=cat)
        ex.cell(row=r, column=3, value=desc); ex.cell(row=r, column=4, value=amt)
dv = DataValidation(type="list", formula1="='Setup'!$A$2:$A$21", allow_blank=True)
ex.add_data_validation(dv); dv.add(f"B2:B{EXP_N}")

# ================= BUDGET =================
bd = wb.create_sheet("Budget"); bd.sheet_view.showGridLines = False
for col, w in zip("ABCDE", [22, 14, 14, 14, 8]):
    bd.column_dimensions[col].width = w
bd["A1"] = "BUDGET"; bd["A1"].font = TITLE_FONT
bd["A3"] = "Monthly income:"; bd["A3"].font = BOLD
inc = bd["D3"]; inc.value = 1400; inc.fill = INPUT_FILL; inc.font = Font(name=ARIAL, bold=True, color="2563EB")
inc.border = BORDER; inc.number_format = MONEY; inc.alignment = Alignment(horizontal="center")

CAT_START = 6
hdr(bd, 5, ["Category", "Budget", "Spent", "Remaining", "key"])
bd.cell(row=5, column=5).font = Font(name=ARIAL, size=8, color="9CA3AF")  # dim the key header
sample_budget = {"Rent / Mortgage":650,"Utilities":100,"Groceries":140,"Dining Out":60,"Transport":40,
                 "Fuel":80,"Insurance":50,"Subscriptions":20,"Shopping":60,"Health":40,
                 "Entertainment":40,"Savings":100,"Other":30}
catN = len(categories)
for i in range(catN):
    r = CAT_START + i
    a = bd.cell(row=r, column=1, value=f"='Setup'!A{2+i}"); a.font = BODY; a.border = BORDER
    b = bd.cell(row=r, column=2, value=sample_budget.get(categories[i])); inp(b); b.number_format = MONEY
    c = bd.cell(row=r, column=3,
        value=f"=IF($A{r}=\"\",\"\",SUMIF(Expenses!$B$2:$B${EXP_N},$A{r},Expenses!$D$2:$D${EXP_N}))")
    d = bd.cell(row=r, column=4, value=f'=IF($A{r}="","",N($B{r})-$C{r})')
    k = bd.cell(row=r, column=5, value=f'=IF($C{r}>0,$C{r}+ROW()/100000,"")')
    for cc in (c, d): cc.fill = CALC_FILL; cc.font = BODY; cc.border = BORDER; cc.number_format = MONEY; cc.alignment = Alignment(horizontal="center")
    k.fill = CALC_FILL; k.font = Font(name=ARIAL, size=8, color="9CA3AF"); k.border = BORDER
CAT_END = CAT_START + catN - 1
TOT = CAT_END + 1
bd.cell(row=TOT, column=1, value="TOTAL").font = BOLD
for col in (2, 3):
    t = bd.cell(row=TOT, column=col, value=f"=SUM({get_column_letter(col)}{CAT_START}:{get_column_letter(col)}{CAT_END})")
    t.font = BOLD; t.border = BORDER; t.number_format = MONEY; t.alignment = Alignment(horizontal="center")
td = bd.cell(row=TOT, column=4, value=f"=B{TOT}-C{TOT}"); td.font = BOLD; td.border = BORDER; td.number_format = MONEY; td.alignment = Alignment(horizontal="center")

# savings goals
SAV_H = TOT + 3
bd.cell(row=SAV_H - 1, column=1, value="SAVINGS GOALS").font = ACCENT
hdr(bd, SAV_H, ["Goal", "Target", "Saved", "% Done"])
sample_sav = [("Emergency fund", 3000, 1200), ("New laptop", 1500, 450)]
SAV_START = SAV_H + 1
for i in range(5):
    r = SAV_START + i
    a = bd.cell(row=r, column=1); inp(a); a.alignment = Alignment(horizontal="left")
    b = bd.cell(row=r, column=2); inp(b); b.number_format = MONEY
    c = bd.cell(row=r, column=3); inp(c); c.number_format = MONEY
    if i < len(sample_sav):
        a.value, b.value, c.value = sample_sav[i]
    d = bd.cell(row=r, column=4, value=f'=IF(OR($A{r}="",$B{r}=""),"",$C{r}/$B{r})')
    d.fill = CALC_FILL; d.font = BODY; d.border = BORDER; d.number_format = PCT; d.alignment = Alignment(horizontal="center")
SAV_END = SAV_START + 4
SAV_TOT = SAV_END + 1
bd.cell(row=SAV_TOT, column=1, value="TOTAL").font = BOLD
for col in (2, 3):
    t = bd.cell(row=SAV_TOT, column=col, value=f"=SUM({get_column_letter(col)}{SAV_START}:{get_column_letter(col)}{SAV_END})")
    t.font = BOLD; t.border = BORDER; t.number_format = MONEY; t.alignment = Alignment(horizontal="center")
tp = bd.cell(row=SAV_TOT, column=4, value=f'=IFERROR(C{SAV_TOT}/B{SAV_TOT},"")'); tp.font = BOLD; tp.border = BORDER; tp.number_format = PCT; tp.alignment = Alignment(horizontal="center")

# ================= HABITS =================
hb = wb.create_sheet("Habits"); hb.sheet_view.showGridLines = False
hb.column_dimensions["A"].width = 18
hb["A1"] = "HABITS"; hb["A1"].font = TITLE_FONT
hb["A3"] = "Month:"; hb["A3"].font = BOLD
mcell = hb["B3"]; mcell.value = "August 2026"; inp(mcell); mcell.alignment = Alignment(horizontal="left")
HROW = 5
headers = ["Habit"] + [str(d) for d in range(1, 32)] + ["Done", "% Month"]
hdr(hb, HROW, headers)
for d in range(1, 32):
    hb.column_dimensions[get_column_letter(1 + d)].width = 3.2
hb.column_dimensions[get_column_letter(33)].width = 7
hb.column_dimensions[get_column_letter(34)].width = 8
hb.freeze_panes = "B6"
H_START = HROW + 1
import random
random.seed(7)
for i in range(len(habits)):
    r = H_START + i
    a = hb.cell(row=r, column=1, value=f"='Setup'!B{2+i}"); a.font = BODY; a.border = BORDER
    for d in range(1, 32):
        c = hb.cell(row=r, column=1 + d); inp(c)
        if d <= 15 and random.random() < 0.6:   # sample marks for first half of month
            c.value = "x"
    done = hb.cell(row=r, column=33, value=f"=COUNTA({get_column_letter(2)}{r}:{get_column_letter(32)}{r})")
    pct = hb.cell(row=r, column=34, value=f"=IFERROR($AG{r}/31,0)")
    for cc in (done, pct): cc.fill = CALC_FILL; cc.font = BOLD; cc.border = BORDER; cc.alignment = Alignment(horizontal="center")
    pct.number_format = PCT
H_END = H_START + len(habits) - 1

# ================= DASHBOARD =================
db = wb.create_sheet("Dashboard"); db.sheet_view.showGridLines = False
db.column_dimensions["A"].width = 2
for col, w in zip("BCDEFG", [22, 14, 4, 22, 16, 12]):
    db.column_dimensions[col].width = w
db["B2"] = "LIFE HQ — DASHBOARD"; db["B2"].font = TITLE_FONT
db["B3"] = "Your month at a glance. Everything here fills itself."; db["B3"].font = SUB_FONT

def stat(cell, label, formula, fmt=MONEY, fill=CARD_FILL):
    lc = db[cell]; lc.value = label; lc.font = BOLD
    v = db[cell[0] + str(int(cell[1:]) + 1)]
    v.value = formula; v.font = BIGNUM; v.number_format = fmt; v.fill = fill; v.border = BORDER
    v.alignment = Alignment(horizontal="center")

stat("B5", "Income",        "=Budget!$D$3")
stat("D5", "Spent",         f"=Budget!$C${TOT}")
stat("F5", "Left to spend", f"=IF(Budget!$D$3=\"\",\"\",Budget!$D$3-Budget!$C${TOT})")
stat("B8", "Budget used",   f"=IFERROR(Budget!$C${TOT}/Budget!$B${TOT},0)", fmt=PCT)
stat("D8", "Saved so far",  f"=Budget!$C${SAV_TOT}")
stat("F8", "Savings goal %",f"=Budget!$D${SAV_TOT}", fmt=PCT)

# habits summary
db["B11"] = "HABITS THIS MONTH"; db["B11"].font = ACCENT
stat("B12", "Completion %",
     f"=IFERROR(SUM(Habits!$AG${H_START}:$AG${H_END})/(COUNTA(Habits!$A${H_START}:$A${H_END})*31),0)", fmt=PCT)
stat("D12", "Check-ins",  f"=SUM(Habits!$AG${H_START}:$AG${H_END})", fmt="#,##0")
lc = db["F12"]; lc.value = "Best habit"; lc.font = BOLD
bh = db["F13"]; bh.value = (f"=IFERROR(INDEX(Habits!$A${H_START}:$A${H_END},"
                            f"MATCH(MAX(Habits!$AG${H_START}:$AG${H_END}),Habits!$AG${H_START}:$AG${H_END},0)),\"-\")")
bh.font = Font(name=ARIAL, bold=True, size=12, color="059669"); bh.fill = CARD_FILL; bh.border = BORDER
bh.alignment = Alignment(horizontal="center")

# top spending categories
db["B16"] = "TOP SPENDING CATEGORIES"; db["B16"].font = ACCENT
hdr(db, 17, ["#", "Category", "Spent"], start=2)  # B,C,D
for k in range(1, 4):
    r = 17 + k
    db.cell(row=r, column=2, value=k).font = BOLD
    key = f"LARGE(Budget!$E${CAT_START}:$E${CAT_END},{k})"
    nm = db.cell(row=r, column=3,
        value=f'=IFERROR(INDEX(Budget!$A${CAT_START}:$A${CAT_END},MATCH({key},Budget!$E${CAT_START}:$E${CAT_END},0)),"—")')
    sp = db.cell(row=r, column=4,
        value=f'=IFERROR(INDEX(Budget!$C${CAT_START}:$C${CAT_END},MATCH({key},Budget!$E${CAT_START}:$E${CAT_END},0)),"")')
    nm.font = BODY; sp.font = BOLD; sp.number_format = MONEY
    for c in (db.cell(row=r, column=2), nm, sp): c.border = BORDER
    sp.alignment = Alignment(horizontal="center")

out = "/home/user/Job-intelligence-agent/products/LifeHQ_Budget_Habit_Tracker.xlsx"
wb.save(out)
print("saved", out, "| TOT row", TOT, "| SAV_TOT", SAV_TOT, "| habits", H_START, H_END)
