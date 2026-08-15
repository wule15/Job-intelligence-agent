import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

wb = openpyxl.Workbook()

# ---------- palette ----------
ARIAL      = "Arial"
HDR_FILL   = PatternFill("solid", fgColor="1F2937")   # slate
HDR_FONT   = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
INPUT_FILL = PatternFill("solid", fgColor="EAF2FB")   # light blue = type here
CALC_FILL  = PatternFill("solid", fgColor="F3F4F6")   # light grey = auto
TITLE_FONT = Font(name=ARIAL, bold=True, size=18, color="111827")
SUB_FONT   = Font(name=ARIAL, size=11, color="4B5563")
BOLD       = Font(name=ARIAL, bold=True, size=11, color="111827")
BODY       = Font(name=ARIAL, size=11, color="111827")
ACCENT     = Font(name=ARIAL, bold=True, size=12, color="2563EB")
thin = Side(style="thin", color="D1D5DB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, headers, start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start+i, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

# ================= READ ME =================
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 100
ws["B2"] = "OVERLOAD"; ws["B2"].font = TITLE_FONT
ws["B3"] = "The gym tracker that tells you what to lift next."; ws["B3"].font = SUB_FONT
lines = [
    ("", ""),
    ("HOW IT WORKS", "h"),
    ("1.  Open the 'Workout Log' tab. Every time you train, add one row per set.", "b"),
    ("2.  Fill only the BLUE cells: Date, Exercise, Weight, Reps, RPE (optional).", "b"),
    ("3.  The grey cells fill themselves — Volume, Estimated 1-Rep-Max, and a 🏆 PR flag when you beat your best.", "b"),
    ("4.  Open the 'Progress' tab to see your best lifts and your suggested next top-set weight for every exercise.", "b"),
    ("", ""),
    ("THE THREE NUMBERS IT CALCULATES FOR YOU", "h"),
    ("•  Estimated 1RM  —  your true strength on a lift, using the Epley formula: weight × (1 + reps/30).", "b"),
    ("•  New PR?  —  automatically flags the set as a personal record when its estimated 1RM beats every past set of that lift.", "b"),
    ("•  Suggested Next Weight  —  your heaviest set + 2.5 kg, so you always know the next target. That is progressive overload.", "b"),
    ("", ""),
    ("TIPS", "h"),
    ("•  Add or rename your lifts on the 'Exercises' tab — the dropdown and the Progress dashboard update automatically.", "b"),
    ("•  Using pounds? Just treat every 'kg' as 'lb' and change the +2.5 step on the Progress tab to +5.", "b"),
    ("•  Keep it in Google Sheets (File ▸ Import) or Excel — it works in both.", "b"),
    ("", ""),
    ("LEGEND", "h"),
    ("BLUE cell = you type here        GREY cell = calculated for you, don't edit", "b"),
]
r = 5
for text, kind in lines:
    c = ws.cell(row=r, column=2, value=text)
    if kind == "h": c.font = ACCENT
    else: c.font = BODY
    r += 1

# ================= EXERCISES =================
ex = wb.create_sheet("Exercises")
ex.sheet_view.showGridLines = False
ex.column_dimensions["A"].width = 30
hdr(ex, 1, ["Exercises (edit freely)"])
exercises = ["Bench Press", "Squat", "Deadlift", "Overhead Press", "Barbell Row",
             "Pull Up", "Incline Bench", "Romanian Deadlift", "Leg Press", "Lat Pulldown"]
for i, name in enumerate(exercises):
    c = ex.cell(row=2+i, column=1, value=name)
    c.font = BODY; c.fill = INPUT_FILL; c.border = BORDER

# ================= WORKOUT LOG =================
log = wb.create_sheet("Workout Log")
log.sheet_view.showGridLines = False
headers = ["Date", "Exercise", "Weight (kg)", "Reps", "RPE", "Volume (kg)", "Est. 1RM (kg)", "New PR?"]
widths  = [13, 20, 12, 8, 8, 12, 14, 11]
for i, w in enumerate(widths):
    log.column_dimensions[chr(65+i)].width = w
hdr(log, 1, headers)
log.freeze_panes = "A2"

# example data: (date, exercise, weight, reps, rpe)
sample = [
    (date(2026,8,4),  "Bench Press",    60,   8, 7),
    (date(2026,8,4),  "Bench Press",    55,   10, 8),
    (date(2026,8,4),  "Overhead Press", 40,   8, 8),
    (date(2026,8,5),  "Squat",          100,  5, 7),
    (date(2026,8,5),  "Barbell Row",    70,   8, 7),
    (date(2026,8,6),  "Deadlift",       120,  5, 8),
    (date(2026,8,11), "Bench Press",    62.5, 8, 8),
    (date(2026,8,11), "Overhead Press", 40,   7, 7),
    (date(2026,8,12), "Squat",          90,   8, 7),
    (date(2026,8,12), "Squat",          100,  6, 8),
    (date(2026,8,12), "Barbell Row",    72.5, 8, 8),
    (date(2026,8,13), "Deadlift",       125,  5, 8),
    (date(2026,8,14), "Bench Press",    62.5, 9, 9),
]
N_INPUT = len(sample)
LAST = 60  # formatted rows for future entries

for idx in range(LAST):
    r = 2 + idx
    # input cells
    for col in range(1, 6):
        c = log.cell(row=r, column=col)
        c.fill = INPUT_FILL; c.font = BODY; c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    if idx < N_INPUT:
        d, exn, wt, rp, rpe = sample[idx]
        log.cell(row=r, column=1, value=d).number_format = "yyyy-mm-dd"
        log.cell(row=r, column=2, value=exn)
        log.cell(row=r, column=3, value=wt)
        log.cell(row=r, column=4, value=rp)
        log.cell(row=r, column=5, value=rpe)
    else:
        log.cell(row=r, column=1).number_format = "yyyy-mm-dd"
    # calc cells
    f = log.cell(row=r, column=6, value=f'=IF($C{r}="","",$C{r}*$D{r})')
    g = log.cell(row=r, column=7, value=f'=IF($C{r}="","",ROUND($C{r}*(1+$D{r}/30),1))')
    h = log.cell(row=r, column=8,
                 value=f'=IF($C{r}="","",IF($G{r}=_xlfn.MAXIFS($G$2:$G{r},$B$2:$B{r},$B{r}),"🏆 PR",""))')
    for c in (f, g, h):
        c.fill = CALC_FILL; c.font = BODY; c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    h.font = Font(name=ARIAL, bold=True, size=11, color="B45309")

# dropdown for Exercise column
dv = DataValidation(type="list", formula1="='Exercises'!$A$2:$A$21", allow_blank=True)
log.add_data_validation(dv)
dv.add(f"B2:B{LAST+1}")

# ================= PROGRESS =================
pr = wb.create_sheet("Progress")
pr.sheet_view.showGridLines = False
pr.column_dimensions["A"].width = 2
for col, w in zip("BCDEFG", [22, 15, 16, 12, 15, 20]):
    pr.column_dimensions[col].width = w

pr["B2"] = "PROGRESS DASHBOARD"; pr["B2"].font = TITLE_FONT
pr["B3"] = "Best lifts, total work, and your next target — updates as you log."; pr["B3"].font = SUB_FONT

pr["B5"] = "This week's total volume (kg):"; pr["B5"].font = BOLD
wk = pr["D5"]
wk.value = '=SUMIFS(\'Workout Log\'!$F$2:$F$61,\'Workout Log\'!$A$2:$A$61,">="&(TODAY()-7))'
wk.font = ACCENT; wk.alignment = Alignment(horizontal="center"); wk.number_format = "#,##0"

ph = ["Exercise", "Best Est. 1RM", "Heaviest Set (kg)", "Total Sets", "Total Volume", "Suggested Next Weight"]
hdr(pr, 7, ph, start=2)
for i in range(len(exercises)):
    r = 8 + i
    a = pr.cell(row=r, column=2, value=f"='Exercises'!A{2+i}")
    a.font = BOLD; a.border = BORDER
    # best est 1RM
    b = pr.cell(row=r, column=3,
        value=(f'=IF(COUNTIF(\'Workout Log\'!$B$2:$B$61,$B{r})=0,"-",'
               f'_xlfn.MAXIFS(\'Workout Log\'!$G$2:$G$61,\'Workout Log\'!$B$2:$B$61,$B{r}))'))
    # heaviest set weight
    c = pr.cell(row=r, column=4,
        value=(f'=IF(COUNTIF(\'Workout Log\'!$B$2:$B$61,$B{r})=0,"-",'
               f'_xlfn.MAXIFS(\'Workout Log\'!$C$2:$C$61,\'Workout Log\'!$B$2:$B$61,$B{r}))'))
    # total sets
    d = pr.cell(row=r, column=5, value=f'=COUNTIF(\'Workout Log\'!$B$2:$B$61,$B{r})')
    # total volume
    e = pr.cell(row=r, column=6,
        value=f'=SUMIF(\'Workout Log\'!$B$2:$B$61,$B{r},\'Workout Log\'!$F$2:$F$61)')
    # suggested next weight = heaviest + 2.5
    fnc = pr.cell(row=r, column=7,
        value=(f'=IF(COUNTIF(\'Workout Log\'!$B$2:$B$61,$B{r})=0,"-",'
               f'_xlfn.MAXIFS(\'Workout Log\'!$C$2:$C$61,\'Workout Log\'!$B$2:$B$61,$B{r})+2.5)'))
    for cell in (b, c, d, e, fnc):
        cell.font = BODY; cell.border = BORDER; cell.alignment = Alignment(horizontal="center")
    fnc.font = Font(name=ARIAL, bold=True, size=11, color="2563EB")
    fnc.number_format = "#,##0.0"
    e.number_format = "#,##0"

out = "/home/user/Job-intelligence-agent/products/OVERLOAD_Gym_Tracker.xlsx"
wb.save(out)
print("saved", out)
