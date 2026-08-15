import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

# ---------- shared palette ----------
ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F2937")
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
INPUT_FILL = PatternFill("solid", fgColor="EAF2FB")
CALC_FILL = PatternFill("solid", fgColor="F3F4F6")
GOLD_FILL = PatternFill("solid", fgColor="FEF3C7")
TITLE_FONT = Font(name=ARIAL, bold=True, size=18, color="111827")
SUB_FONT = Font(name=ARIAL, size=11, color="4B5563")
BOLD = Font(name=ARIAL, bold=True, size=11, color="111827")
BODY = Font(name=ARIAL, size=11, color="111827")
thin = Side(style="thin", color="D1D5DB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, headers, accent, start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start + i, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def inp(c):
    c.fill = INPUT_FILL; c.font = BODY; c.border = BORDER
    c.alignment = Alignment(horizontal="center")

def upgrade_tab(wb, product, accent_hex, tagline, pro_features, price):
    up = wb.create_sheet("★ Upgrade")
    up.sheet_view.showGridLines = False
    up.column_dimensions["A"].width = 2
    up.column_dimensions["B"].width = 95
    ACC = Font(name=ARIAL, bold=True, size=13, color=accent_hex)
    up["B2"] = f"You're using {product} — Free Lite"; up["B2"].font = TITLE_FONT
    up["B3"] = tagline; up["B3"].font = SUB_FONT
    up["B5"] = "The full version adds:"; up["B5"].font = ACC
    r = 6
    for feat in pro_features:
        c = up.cell(row=r, column=2, value=f"✓  {feat}"); c.font = BODY; r += 1
    r += 1
    box = up.cell(row=r, column=2, value=f"👉  Get {product} (full version) — {price}")
    box.font = Font(name=ARIAL, bold=True, size=13, color="111827")
    box.fill = GOLD_FILL; box.border = BORDER
    box.alignment = Alignment(horizontal="center", vertical="center")
    up.row_dimensions[r].height = 30
    link = up.cell(row=r + 1, column=2, value="https://[your-store-link]   ← replace with your Gumroad/Etsy URL")
    link.font = Font(name=ARIAL, size=11, color="2563EB")
    up.cell(row=r + 3, column=2,
            value="Free to use and share. If it helped you, the full version is a few dollars and does a lot more.").font = SUB_FONT
    return up

def readme(wb, title, tagline, steps, accent_hex):
    ws = wb.active; ws.title = "Read Me"; ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2; ws.column_dimensions["B"].width = 95
    ACC = Font(name=ARIAL, bold=True, size=12, color=accent_hex)
    ws["B2"] = title; ws["B2"].font = TITLE_FONT
    ws["B3"] = tagline; ws["B3"].font = SUB_FONT
    ws["B5"] = "HOW TO USE"; ws["B5"].font = ACC
    r = 6
    for s in steps:
        ws.cell(row=r, column=2, value=s).font = BODY; r += 1
    ws.cell(row=r + 1, column=2, value="BLUE = you fill in     GREY = calculated for you").font = SUB_FONT
    ws.cell(row=r + 2, column=2, value="This is the free Lite version — see the ★ Upgrade tab for the full one.").font = ACC
    return ws


# ============================================================
# 1) OVERLOAD LITE — gym
# ============================================================
def build_overload_lite():
    wb = openpyxl.Workbook()
    ACC = "2563EB"
    readme(wb, "OVERLOAD — Free Lite", "Log a set; it calculates your 1-rep-max and flags your PRs.",
           ["1.  Go to the 'Workout Log' tab.",
            "2.  Fill the blue cells: Date, Exercise, Weight, Reps.",
            "3.  Volume, Estimated 1RM and a 🏆 PR flag fill in automatically."], ACC)
    log = wb.create_sheet("Workout Log"); log.sheet_view.showGridLines = False
    heads = ["Date", "Exercise", "Weight", "Reps", "Volume", "Est. 1RM", "New PR?"]
    for i, w in enumerate([13, 20, 10, 8, 10, 12, 11]):
        log.column_dimensions[chr(65 + i)].width = w
    hdr(log, 1, heads, ACC); log.freeze_panes = "A2"
    sample = [(date(2026, 8, 4), "Bench Press", 60, 8), (date(2026, 8, 4), "Bench Press", 55, 10),
              (date(2026, 8, 6), "Squat", 100, 5), (date(2026, 8, 11), "Bench Press", 62.5, 8),
              (date(2026, 8, 13), "Deadlift", 120, 5)]
    for idx in range(40):
        r = 2 + idx
        for col in range(1, 5):
            inp(log.cell(row=r, column=col))
        log.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        if idx < len(sample):
            d, e, wt, rp = sample[idx]
            log.cell(row=r, column=1, value=d); log.cell(row=r, column=2, value=e)
            log.cell(row=r, column=3, value=wt); log.cell(row=r, column=4, value=rp)
        f = log.cell(row=r, column=5, value=f'=IF(AND(ISNUMBER($C{r}),ISNUMBER($D{r})),$C{r}*$D{r},"")')
        g = log.cell(row=r, column=6, value=f'=IF(AND(ISNUMBER($C{r}),ISNUMBER($D{r})),ROUND($C{r}*(1+$D{r}/30),1),"")')
        h = log.cell(row=r, column=7, value=(
            f'=IF($F{r}="","",IF(ROUND($C{r}*(1+$D{r}/30),1)='
            f'ROUND(SUMPRODUCT(MAX(($B$2:$B{r}=$B{r})*N($C$2:$C{r})*(1+N($D$2:$D{r})/30))),1),"🏆 PR",""))'))
        for c in (f, g, h):
            c.fill = CALC_FILL; c.font = BODY; c.border = BORDER; c.alignment = Alignment(horizontal="center")
        h.font = Font(name=ARIAL, bold=True, color="B45309")
    upgrade_tab(wb, "OVERLOAD", ACC,
                "Free to log & spot PRs. The full version turns it into a coach.",
                ["A Progress dashboard — best 1RM, total volume, weekly workload",
                 "Suggested NEXT weight for every lift (progressive overload, done for you)",
                 "Up to 25 custom lifts with a dropdown",
                 "Best-lift tracking and kg⇄lb switch in one cell"], "$5")
    wb.save("/home/user/Job-intelligence-agent/products/OVERLOAD_Lite_Gym_Tracker.xlsx")
    print("saved OVERLOAD Lite")


# ============================================================
# 2) BACKLOG BOSS LITE — gaming
# ============================================================
def build_backlog_lite():
    wb = openpyxl.Workbook()
    ACC = "7C3AED"
    readme(wb, "BACKLOG BOSS — Free Lite", "Log your games, track status, see cost-per-hour.",
           ["1.  Go to the 'Game Library' tab.",
            "2.  Add a game per row; pick a Status from the dropdown; add hours and price.",
            "3.  Cost/Hour fills in automatically. The counters up top update live."], ACC)
    gl = wb.create_sheet("Game Library"); gl.sheet_view.showGridLines = False
    for i, w in enumerate([26, 16, 13, 12, 11, 11]):
        gl.column_dimensions[chr(65 + i)].width = w
    # teaser counters
    gl["H1"] = "Games:"; gl["H1"].font = BOLD
    gl["I1"] = "=COUNTA($A$4:$A$103)"; gl["I1"].font = Font(name=ARIAL, bold=True, color=ACC)
    gl["H2"] = "Backlog:"; gl["H2"].font = BOLD
    gl["I2"] = '=COUNTIF($C$4:$C$103,"Backlog")'; gl["I2"].font = Font(name=ARIAL, bold=True, color=ACC)
    hdr(gl, 3, ["Title", "Platform", "Status", "Hours", "Price", "Cost/Hour"], ACC)
    gl.freeze_panes = "A4"
    sample = [("Elden Ring", "PC", "Beaten", 95, 59.99), ("Baldur's Gate 3", "PC", "Playing", 60, 59.99),
              ("God of War", "PS5", "Backlog", None, 49.99), ("Stardew Valley", "Switch", "Playing", 120, 13.99),
              ("Hollow Knight", "PC", "Backlog", None, 14.99)]
    for idx in range(100):
        r = 4 + idx
        for col in range(1, 6):
            inp(gl.cell(row=r, column=col))
        gl.cell(row=r, column=5).number_format = "#,##0.00"
        if idx < len(sample):
            t, p, s, h, pr = sample[idx]
            gl.cell(row=r, column=1, value=t); gl.cell(row=r, column=2, value=p); gl.cell(row=r, column=3, value=s)
            if h is not None: gl.cell(row=r, column=4, value=h)
            gl.cell(row=r, column=5, value=pr)
        ch = gl.cell(row=r, column=6, value=f'=IF(AND(ISNUMBER($D{r}),ISNUMBER($E{r}),$D{r}>0),ROUND($E{r}/$D{r},2),"")')
        ch.fill = CALC_FILL; ch.font = BODY; ch.border = BORDER
        ch.alignment = Alignment(horizontal="center"); ch.number_format = "#,##0.00"
    dv = DataValidation(type="list", formula1='"Backlog,Playing,Beaten,Completed,Dropped,Wishlist"', allow_blank=True)
    gl.add_data_validation(dv); dv.add("C4:C103")
    upgrade_tab(wb, "Backlog Boss", ACC,
                "Free to track your library. The full version makes it an app.",
                ["A live Dashboard — completion %, hours, spend, Pile of Shame counter",
                 "Best-Value leaderboard ranking your games by cost-per-hour",
                 "'What To Play Next' random backlog picker",
                 "Genre, ownership, rating and personal review fields; 120 games"], "$4.99")
    wb.save("/home/user/Job-intelligence-agent/products/BacklogBoss_Lite_Game_Tracker.xlsx")
    print("saved Backlog Boss Lite")


# ============================================================
# 3) LIFE HQ LITE — budget only
# ============================================================
def build_lifehq_lite():
    wb = openpyxl.Workbook()
    ACC = "059669"
    readme(wb, "LIFE HQ — Free Lite (Budget)", "A simple monthly budget: set it, fill spent, see what's left.",
           ["1.  Go to the 'Budget' tab.",
            "2.  Set a Budget and type what you've Spent for each category.",
            "3.  Remaining and the % used total calculate automatically."], ACC)
    bd = wb.create_sheet("Budget"); bd.sheet_view.showGridLines = False
    for i, w in enumerate([22, 14, 14, 14]):
        bd.column_dimensions[chr(65 + i)].width = w
    bd["A1"] = "MONTHLY BUDGET"; bd["A1"].font = TITLE_FONT
    cats = ["Rent / Mortgage", "Utilities", "Groceries", "Dining Out", "Transport", "Fuel",
            "Subscriptions", "Shopping", "Health", "Entertainment", "Savings", "Other"]
    START = 4
    hdr(bd, 3, ["Category", "Budget", "Spent", "Remaining"], ACC)
    demo_budget = {"Rent / Mortgage": 650, "Groceries": 140, "Utilities": 100}
    demo_spent = {"Rent / Mortgage": 650, "Groceries": 110, "Utilities": 90}
    for i, cat in enumerate(cats):
        r = START + i
        a = bd.cell(row=r, column=1, value=cat); a.font = BODY; a.border = BORDER
        b = bd.cell(row=r, column=2, value=demo_budget.get(cat)); inp(b); b.number_format = "#,##0.00"
        c = bd.cell(row=r, column=3, value=demo_spent.get(cat)); inp(c); c.number_format = "#,##0.00"
        d = bd.cell(row=r, column=4, value=f'=IF($A{r}="","",N($B{r})-N($C{r}))')
        d.fill = CALC_FILL; d.font = BODY; d.border = BORDER; d.number_format = "#,##0.00"; d.alignment = Alignment(horizontal="center")
    END = START + len(cats) - 1
    TOT = END + 1
    bd.cell(row=TOT, column=1, value="TOTAL").font = BOLD
    for col, L in ((2, "B"), (3, "C")):
        t = bd.cell(row=TOT, column=col, value=f"=SUM({L}{START}:{L}{END})")
        t.font = BOLD; t.border = BORDER; t.number_format = "#,##0.00"; t.alignment = Alignment(horizontal="center")
    td = bd.cell(row=TOT, column=4, value=f"=B{TOT}-C{TOT}"); td.font = BOLD; td.border = BORDER
    td.number_format = "#,##0.00"; td.alignment = Alignment(horizontal="center")
    bd.cell(row=TOT + 2, column=1, value="% of budget used:").font = BOLD
    pu = bd.cell(row=TOT + 2, column=2, value=f"=IFERROR(C{TOT}/B{TOT},0)")
    pu.font = Font(name=ARIAL, bold=True, size=14, color=ACC); pu.number_format = "0%"; pu.alignment = Alignment(horizontal="center")
    upgrade_tab(wb, "LIFE HQ", ACC,
                "Free budget basics. The full version runs your whole month.",
                ["Auto expense log with dropdown categories (spending flows in by itself)",
                 "Savings goals with automatic % complete",
                 "A full 31-day habit tracker with streaks and completion %",
                 "One live Dashboard combining money + habits, plus top-spending ranking"], "$12")
    wb.save("/home/user/Job-intelligence-agent/products/LifeHQ_Lite_Budget_Tracker.xlsx")
    print("saved LIFE HQ Lite")


build_overload_lite()
build_backlog_lite()
build_lifehq_lite()
print("done")
