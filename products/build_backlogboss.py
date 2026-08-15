import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

wb = openpyxl.Workbook()

# ---------- palette (shared brand with OVERLOAD) ----------
ARIAL      = "Arial"
HDR_FILL   = PatternFill("solid", fgColor="1F2937")
HDR_FONT   = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
INPUT_FILL = PatternFill("solid", fgColor="EAF2FB")
CALC_FILL  = PatternFill("solid", fgColor="F3F4F6")
CARD_FILL  = PatternFill("solid", fgColor="EDE9FE")   # soft purple = gaming accent
SHAME_FILL = PatternFill("solid", fgColor="FEF3C7")
TITLE_FONT = Font(name=ARIAL, bold=True, size=18, color="111827")
SUB_FONT   = Font(name=ARIAL, size=11, color="4B5563")
BOLD       = Font(name=ARIAL, bold=True, size=11, color="111827")
BODY       = Font(name=ARIAL, size=11, color="111827")
ACCENT     = Font(name=ARIAL, bold=True, size=12, color="7C3AED")
BIGNUM     = Font(name=ARIAL, bold=True, size=16, color="7C3AED")
thin = Side(style="thin", color="D1D5DB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, headers, start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start+i, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

N = 121  # library rows 2..121 (120 games)
GL = "'Game Library'"

# ================= READ ME =================
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 100
ws["B2"] = "BACKLOG BOSS"; ws["B2"].font = TITLE_FONT
ws["B3"] = "Beat your pile of shame. Your whole game library, owned by you."; ws["B3"].font = SUB_FONT
lines = [
    ("", ""),
    ("HOW IT WORKS", "h"),
    ("1.  Go to the 'Game Library' tab. Add one row per game — fill the BLUE cells.", "b"),
    ("2.  Pick Platform, Genre, Status and Ownership from the dropdowns. Add hours, price paid and a rating as you go.", "b"),
    ("3.  The grey cells fill themselves (Cost/Hour). The 'Dashboard' tab updates live — counts, totals, best-value ranking.", "b"),
    ("4.  Can't decide what to play? The Dashboard's 'What To Play Next' picks a random game from your backlog.", "b"),
    ("", ""),
    ("WHAT THE DASHBOARD SHOWS YOU", "h"),
    ("•  Your collection at a glance — total games, hours, money spent, average rating, completion %.", "b"),
    ("•  A live count of every status, plus your Pile of Shame (games sitting in the backlog).", "b"),
    ("•  Best Value leaderboard — the games that gave you the most play per dollar, ranked automatically.", "b"),
    ("•  What To Play Next — a random backlog pick to kill decision paralysis.", "b"),
    ("", ""),
    ("TIPS", "h"),
    ("•  Add or edit Platforms, Genres and Statuses on the 'Lists' tab — the dropdowns update automatically.", "b"),
    ("•  Price Paid has no currency symbol, so it works in any currency — just be consistent.", "b"),
    ("•  Reroll 'What To Play Next': click any empty cell and press Enter (or Delete) to recalculate.", "b"),
    ("•  Honest note: a spreadsheet can't auto-import from Steam/PlayStation. Entry is manual — that's the point: it's yours, private, no login, and fully customisable.", "b"),
    ("", ""),
    ("LEGEND", "h"),
    ("BLUE cell = you type/pick here        GREY cell = calculated for you, don't edit", "b"),
]
r = 5
for text, kind in lines:
    c = ws.cell(row=r, column=2, value=text)
    c.font = ACCENT if kind == "h" else BODY
    r += 1

# ================= LISTS =================
lst = wb.create_sheet("Lists")
lst.sheet_view.showGridLines = False
for col, w in zip("ABCD", [18, 16, 14, 14]):
    lst.column_dimensions[col].width = w
hdr(lst, 1, ["Platforms", "Genres", "Statuses", "Ownership"])
platforms = ["PC","PlayStation 5","PlayStation 4","Xbox Series","Xbox One",
             "Nintendo Switch","Steam Deck","Mobile","Retro","Other"]
genres = ["Action","Adventure","RPG","JRPG","Shooter","Strategy","Simulation","Sports",
          "Racing","Puzzle","Platformer","Metroidvania","Roguelike","Horror","Fighting","MMO","Indie","Other"]
statuses = ["Backlog","Playing","Beaten","Completed","Dropped","Wishlist"]
ownership = ["Physical","Digital","Subscription"]
def fill_col(col, items):
    for i, v in enumerate(items):
        c = lst.cell(row=2+i, column=col, value=v)
        c.font = BODY; c.fill = INPUT_FILL; c.border = BORDER
fill_col(1, platforms); fill_col(2, genres); fill_col(3, statuses); fill_col(4, ownership)

# ================= GAME LIBRARY =================
gl = wb.create_sheet("Game Library")
gl.sheet_view.showGridLines = False
headers = ["Title","Platform","Genre","Status","Ownership","Rating /10","Hours Played",
           "Price Paid","Cost / Hour","Date Beaten","Notes / Review","(auto)","(auto)"]
widths  = [26,15,14,12,13,10,12,11,11,13,30,7,7]
for i, w in enumerate(widths):
    gl.column_dimensions[chr(65+i)].width = w
hdr(gl, 1, headers)
gl.freeze_panes = "A2"

sample = [
    # title, platform, genre, status, ownership, rating, hours, price, date_beaten, notes
    ("Elden Ring","PC","Roguelike","Beaten","Digital",10,95,59.99,date(2026,2,10),"Masterpiece"),
    ("Baldur's Gate 3","PC","RPG","Playing","Digital",10,60,59.99,None,"Act 2, obsessed"),
    ("Hades","PC","Roguelike","Completed","Digital",9,40,24.99,date(2025,12,1),"One more run..."),
    ("Cyberpunk 2077","PC","RPG","Beaten","Digital",8,55,29.99,date(2026,1,15),"Great after patches"),
    ("Stardew Valley","Nintendo Switch","Simulation","Playing","Digital",9,120,13.99,None,"Cozy forever"),
    ("God of War Ragnarok","PlayStation 5","Action","Backlog","Physical",None,None,49.99,None,""),
    ("Zelda: Tears of the Kingdom","Nintendo Switch","Adventure","Backlog","Physical",None,None,69.99,None,""),
    ("Red Dead Redemption 2","PC","Action","Backlog","Digital",None,None,39.99,None,""),
    ("Celeste","PC","Platformer","Beaten","Digital",9,12,19.99,date(2025,11,20),"Hard but fair"),
    ("Hollow Knight","PC","Metroidvania","Backlog","Digital",None,None,14.99,None,""),
    ("Starfield","PC","RPG","Dropped","Digital",6,25,69.99,None,"Lost interest"),
    ("Hollow Knight: Silksong","PC","Metroidvania","Wishlist",None,None,None,None,None,"Someday..."),
    ("Persona 5 Royal","PlayStation 5","JRPG","Backlog","Physical",None,None,59.99,None,""),
]
N_INPUT = len(sample)
for idx in range(N - 1):
    r = 2 + idx
    for col in range(1, 12):  # A-K input area
        c = gl.cell(row=r, column=col)
        c.fill = INPUT_FILL; c.font = BODY; c.border = BORDER
        if col not in (1, 11):
            c.alignment = Alignment(horizontal="center")
    gl.cell(row=r, column=9).fill = CALC_FILL   # Cost/Hour is calculated
    if idx < N_INPUT:
        t,pf,gn,st,ow,rt,hr,pc,db,nt = sample[idx]
        gl.cell(row=r, column=1, value=t)
        gl.cell(row=r, column=2, value=pf)
        gl.cell(row=r, column=3, value=gn)
        gl.cell(row=r, column=4, value=st)
        gl.cell(row=r, column=5, value=ow)
        if rt is not None: gl.cell(row=r, column=6, value=rt)
        if hr is not None: gl.cell(row=r, column=7, value=hr)
        if pc is not None: gl.cell(row=r, column=8, value=pc)
        if db is not None: gl.cell(row=r, column=10, value=db).number_format = "yyyy-mm-dd"
        gl.cell(row=r, column=11, value=nt)
    else:
        gl.cell(row=r, column=10).number_format = "yyyy-mm-dd"
    # Cost/Hour = price / hours, guarded
    ch = gl.cell(row=r, column=9,
        value=f'=IF(AND(ISNUMBER($G{r}),ISNUMBER($H{r}),$G{r}>0),ROUND($H{r}/$G{r},2),"")')
    ch.fill = CALC_FILL; ch.font = BODY; ch.border = BORDER
    ch.alignment = Alignment(horizontal="center"); ch.number_format = "#,##0.00"
    gl.cell(row=r, column=8).number_format = "#,##0.00"
    # helper L: sequential backlog index (for random picker)
    l = gl.cell(row=r, column=12, value=f'=IF($D{r}="Backlog",COUNTIF($D$2:$D{r},"Backlog"),"")')
    # helper M: unique value-key for best-value ranking (lower cost/hr = better; +row keeps keys unique)
    m = gl.cell(row=r, column=13, value=f'=IF(ISNUMBER($I{r}),$I{r}+ROW()/100000,"")')
    for c in (l, m):
        c.fill = CALC_FILL; c.font = Font(name=ARIAL, size=8, color="9CA3AF"); c.border = BORDER
        c.alignment = Alignment(horizontal="center")

# dropdowns
def add_dv(rng, formula):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    gl.add_data_validation(dv); dv.add(rng)
add_dv(f"B2:B{N}", "='Lists'!$A$2:$A$21")
add_dv(f"C2:C{N}", "='Lists'!$B$2:$B$21")
add_dv(f"D2:D{N}", "='Lists'!$C$2:$C$8")
add_dv(f"E2:E{N}", "='Lists'!$D$2:$D$5")

# ================= DASHBOARD =================
db = wb.create_sheet("Dashboard")
db.sheet_view.showGridLines = False
db.column_dimensions["A"].width = 2
for col, w in zip("BCDEFG", [24, 14, 4, 24, 14, 12]):
    db.column_dimensions[col].width = w
db["B2"] = "BACKLOG BOSS — DASHBOARD"; db["B2"].font = TITLE_FONT
db["B3"] = "Live stats from your library. Nothing to fill in here."; db["B3"].font = SUB_FONT

def stat(cell, label, formula, fmt="#,##0", big=True, fill=CARD_FILL):
    lc = db[cell]; lc.value = label; lc.font = BOLD
    vcell = db[cell[0] + str(int(cell[1:]) + 1)]
    vcell.value = formula; vcell.font = BIGNUM if big else BOLD
    vcell.number_format = fmt; vcell.fill = fill; vcell.border = BORDER
    vcell.alignment = Alignment(horizontal="center")

CA   = f'COUNTA({GL}!$A$2:$A${N})'
cnt  = lambda s: f'COUNTIF({GL}!$D$2:$D${N},"{s}")'

# top KPI row
stat("B5",  "Total games",     f'={CA}')
stat("D5",  "Hours played",    f'=SUM({GL}!$G$2:$G${N})')
stat("F5",  "Total spent",     f'=SUM({GL}!$H$2:$H${N})', fmt="#,##0.00")
stat("B8",  "Avg rating /10",  f'=IFERROR(ROUND(AVERAGEIF({GL}!$F$2:$F${N},">0"),1),"-")', fmt="0.0")
stat("D8",  "Completion %",
     f'=IFERROR(({cnt("Beaten")}+{cnt("Completed")})/({CA}-{cnt("Wishlist")}),0)', fmt="0%")
stat("F8",  "Pile of Shame",   f'={cnt("Backlog")}', fill=SHAME_FILL)

# status breakdown
db["B11"] = "LIBRARY BY STATUS"; db["B11"].font = ACCENT
hdr(db, 12, ["Status", "Count"], start=2)
for i, s in enumerate(statuses):
    rr = 13 + i
    a = db.cell(row=rr, column=2, value=s); a.font = BODY; a.border = BORDER
    v = db.cell(row=rr, column=3, value=f'={cnt(s)}'); v.font = BOLD; v.border = BORDER
    v.alignment = Alignment(horizontal="center")

# what to play next
db["E11"] = "WHAT TO PLAY NEXT"; db["E11"].font = ACCENT
pick = db["E12"]
pick.value = (f'=IFERROR(INDEX({GL}!$A$2:$A${N},'
              f'MATCH(RANDBETWEEN(1,{cnt("Backlog")}),{GL}!$L$2:$L${N},0)),"— add backlog games —")')
pick.font = Font(name=ARIAL, bold=True, size=13, color="7C3AED")
pick.fill = CARD_FILL; pick.border = BORDER
pick.alignment = Alignment(horizontal="center", vertical="center")
db.merge_cells("E12:F13")
db["E14"] = "(click an empty cell + Enter to reroll)"; db["E14"].font = Font(name=ARIAL, size=9, italic=True, color="9CA3AF")

# best value leaderboard
db["E16"] = "BEST VALUE  (lowest cost / hour)"; db["E16"].font = ACCENT
hdr(db, 17, ["#", "Game", "Cost/Hr"], start=5)  # E,F,G
key = f'SMALL({GL}!$M$2:$M${N},{{k}})'
for k in range(1, 6):
    rr = 17 + k
    db.cell(row=rr, column=5, value=k).font = BOLD
    kk = key.format(k=k)
    nm = db.cell(row=rr, column=6,
        value=f'=IFERROR(INDEX({GL}!$A$2:$A${N},MATCH({kk},{GL}!$M$2:$M${N},0)),"—")')
    ch = db.cell(row=rr, column=7,
        value=f'=IFERROR(INDEX({GL}!$I$2:$I${N},MATCH({kk},{GL}!$M$2:$M${N},0)),"")')
    nm.font = BODY; ch.font = BOLD; ch.number_format = "#,##0.00"
    for c in (db.cell(row=rr, column=5), nm, ch):
        c.border = BORDER
    ch.alignment = Alignment(horizontal="center")

out = "/home/user/Job-intelligence-agent/products/BacklogBoss_Game_Tracker.xlsx"
wb.save(out)
print("saved", out)
